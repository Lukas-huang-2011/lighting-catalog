"""Fill the order_template.xlsx with product data from the database."""

import io
import os
import datetime
from copy import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "order_template.xlsx")

# Column positions in the template (1-based)
COL_SEQ    = 1   # 序号
COL_BRAND  = 2   # 品牌
COL_CODE   = 3   # 编号
COL_NAME   = 4   # 名称
COL_IMAGE  = 5   # 图片
COL_COLOR  = 6   # 颜色
COL_TYPE   = 7   # 种类
COL_DIM    = 8   # 尺寸 (cm)  ← also receives dimension-drawing image
COL_LIGHT  = 9   # 光源参数
COL_DELIV  = 10  # 到货时间
COL_PRICE  = 11  # 零售单价
COL_QTY    = 12  # 数量
COL_TOTAL  = 13  # 合计  (=Kn*Ln)
COL_DISC   = 14  # 折扣
COL_FINAL  = 15  # 折后价 (=Mn*Nn)

PRODUCT_START_ROW = 9  # First data row in template
ROW_HEIGHT_PX     = 80  # Row height for product rows (used when images present)

# ── Lighting-type keyword → Chinese translation ───────────────────────────────
_TYPE_KW_ZH = [
    ("pendant",    "吊灯"), ("suspension", "吊灯"), ("chandelier", "吊灯"),
    ("hanging",    "吊灯"), ("sospension", "吊灯"),
    ("wall",       "壁灯"), ("sconce",     "壁灯"), ("aplique",    "壁灯"),
    ("table",      "台灯"), ("desk",       "台灯"),
    ("floor",      "落地灯"),
    ("ceiling",    "吸顶灯"), ("flush",    "吸顶灯"), ("plafon",   "吸顶灯"),
    ("spot",       "射灯"),  ("spotlight", "射灯"),
    ("downlight",  "筒灯"),  ("recessed",  "筒灯"),
    ("track",      "轨道灯"),
    ("strip",      "灯带"),  ("linear",    "线条灯"), ("profile",  "线条灯"),
    ("outdoor",    "户外灯"), ("exterior", "户外灯"),
    ("garden",     "庭院灯"), ("street",   "路灯"),
    ("panel",      "面板灯"), ("bollard",  "地埋灯"),
]


def _detect_zh_type(text: str) -> str:
    """Return Chinese lighting-type for the first keyword found in text, or ''."""
    lower = (text or "").lower()
    for kw, zh in _TYPE_KW_ZH:
        if kw in lower:
            return zh
    return ""


def _extract_brand(pdf_name: str) -> str:
    """Turn a PDF filename into a readable brand name."""
    if not pdf_name:
        return ""
    name = pdf_name.replace(".pdf", "").replace(".PDF", "")
    return name.replace("_", " ").replace("-", " ").title()


def _copy_row_style(ws, src_row: int, dst_row: int):
    """Copy cell formatting from one row to another."""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font      = copy(src.font)
            dst.fill      = copy(src.fill)
            dst.border    = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


def _pil_to_xl_image(pil_img: PILImage.Image, max_px: int = 120) -> XLImage:
    """Convert a PIL Image to an openpyxl Image object sized for a cell."""
    w, h = pil_img.size
    scale = min(max_px / w, max_px / h, 1.0)
    thumb = pil_img.resize((int(w * scale), int(h * scale)), PILImage.LANCZOS)
    buf = io.BytesIO()
    thumb.save(buf, format="PNG")
    buf.seek(0)
    return XLImage(buf)


def build_excel_from_template(
    products: list,
    order_info: dict | None = None,
    product_images: dict | None = None,   # {product_index: PIL.Image}  → 图片 column
    dim_images:     dict | None = None,   # {product_index: PIL.Image}  → 尺寸 column
) -> bytes:
    """
    Fill order_template.xlsx with product rows and return the file as bytes.

    order_info keys (all optional):
      order_number, date, customer_name, contact_person, phone

    product_images: optional dict mapping product list index → PIL Image
      to embed in the 图片 (E) column

    dim_images: optional dict mapping product list index → PIL Image
      to embed in the 尺寸 (H) column (dimension drawings with measurement labels)
    """
    if order_info is None:
        order_info = {}
    if product_images is None:
        product_images = {}
    if dim_images is None:
        dim_images = {}

    n = len(products)
    if n == 0:
        with open(TEMPLATE_PATH, "rb") as f:
            return f.read()

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # ── 0. Clear any sample images baked into the template ───────────────────
    ws._images.clear()

    # ── 1. Fill order header ──────────────────────────────────────────────────
    if order_info.get("order_number"):
        ws["C3"] = order_info["order_number"]
    if order_info.get("date"):
        ws["C4"] = order_info["date"]
    else:
        ws["C4"] = datetime.date.today()
    if order_info.get("customer_name"):
        ws["C5"] = order_info["customer_name"]
    if order_info.get("contact_person"):
        ws["C6"] = order_info["contact_person"]
    if order_info.get("phone"):
        ws["E6"] = order_info["phone"]

    # ── 2. Fix merged cells before inserting rows ─────────────────────────────
    # openpyxl does NOT shift merged cell ranges when inserting rows, so footer
    # merges that start at row >= PRODUCT_START_ROW+1 block writes to new rows.
    # Solution: unmerge them, insert rows, then re-merge at shifted positions.
    extra_rows = max(n - 1, 0)
    footer_merges = []   # (min_row, min_col, max_row, max_col)

    if extra_rows > 0:
        to_remove = []
        for merge in list(ws.merged_cells.ranges):
            if merge.min_row >= PRODUCT_START_ROW + 1:
                footer_merges.append((
                    merge.min_row, merge.min_col,
                    merge.max_row, merge.max_col,
                ))
                to_remove.append(str(merge))
        for ref in to_remove:
            ws.merged_cells.remove(ref)

        # Insert blank rows for products 2..n
        ws.insert_rows(PRODUCT_START_ROW + 1, amount=extra_rows)

        # Copy style and row height from template product row to each new row
        src_height = ws.row_dimensions[PRODUCT_START_ROW].height or 146
        for i in range(1, n):
            _copy_row_style(ws, PRODUCT_START_ROW, PRODUCT_START_ROW + i)
            ws.row_dimensions[PRODUCT_START_ROW + i].height = src_height

        # Re-merge footer ranges at shifted positions
        for (r1, c1, r2, c2) in footer_merges:
            ws.merge_cells(
                start_row=r1 + extra_rows, start_column=c1,
                end_row=r2 + extra_rows,   end_column=c2,
            )

    # ── 3. Write product rows ─────────────────────────────────────────────────
    for i, product in enumerate(products):
        row = PRODUCT_START_ROW + i
        codes    = product.get("codes") or []
        ef       = product.get("extra_fields") or {}
        pdf_info = product.get("pdfs") or {}
        brand    = _extract_brand(pdf_info.get("name") or product.get("brand") or "")
        price    = product.get("price")
        qty      = product.get("_qty", 1)
        discount = product.get("_discount", 1)

        # ── 颜色: UI override → PDF field → "如图" (template placeholder) ──────
        if product.get("_color") is not None:
            color = product["_color"] or ""          # explicit UI value (may be empty)
        else:
            color = product.get("color") or "如图"   # auto: PDF value or placeholder

        # ── 到货时间 ─────────────────────────────────────────────────────────
        delivery = product.get("_delivery") if product.get("_delivery") is not None else "现货"

        # ── 种类: UI override → auto-detect from description/name → raw description ──
        if product.get("_category") is not None:
            category = product["_category"] or ""    # explicit UI value
        else:
            raw_desc = product.get("description") or ""
            category = (
                _detect_zh_type(raw_desc)
                or _detect_zh_type(product.get("name") or "")
                or raw_desc
            )

        ws.cell(row=row, column=COL_SEQ).value   = i + 1
        ws.cell(row=row, column=COL_BRAND).value = brand
        ws.cell(row=row, column=COL_CODE).value  = ", ".join(str(c) for c in codes)
        ws.cell(row=row, column=COL_NAME).value  = product.get("name") or ""
        ws.cell(row=row, column=COL_COLOR).value = color
        ws.cell(row=row, column=COL_TYPE).value  = category
        ws.cell(row=row, column=COL_DIM).value   = product.get("dimensions") or ""
        light_parts = [
            product.get("light_source") or ef.get("light_source") or "",
            product.get("cct") or ef.get("cct") or "",
            product.get("wattage") or ef.get("wattage") or "",
        ]
        ws.cell(row=row, column=COL_LIGHT).value = "  ".join(p for p in light_parts if p)
        ws.cell(row=row, column=COL_DELIV).value = delivery
        ws.cell(row=row, column=COL_PRICE).value = price
        ws.cell(row=row, column=COL_QTY).value   = qty
        ws.cell(row=row, column=COL_TOTAL).value = f"=K{row}*L{row}"
        ws.cell(row=row, column=COL_DISC).value  = discount
        ws.cell(row=row, column=COL_FINAL).value = f"=M{row}*N{row}"

        # ── Embed 图片 (product illustration) if provided ─────────────────────
        if i in product_images and product_images[i] is not None:
            try:
                xl_img = _pil_to_xl_image(product_images[i], max_px=185)
                xl_img.anchor = f"{get_column_letter(COL_IMAGE)}{row}"
                ws.add_image(xl_img)
            except Exception:
                pass

        # ── Embed 尺寸 (dimension drawing) if provided ────────────────────────
        if i in dim_images and dim_images[i] is not None:
            try:
                xl_img = _pil_to_xl_image(dim_images[i], max_px=160)
                xl_img.anchor = f"{get_column_letter(COL_DIM)}{row}"
                ws.add_image(xl_img)
            except Exception:
                pass

    # ── 4. Update footer formulas ─────────────────────────────────────────────
    last_prod = PRODUCT_START_ROW + n - 1
    sub_row   = last_prod + 1
    other_row = last_prod + 2
    total_row = last_prod + 3

    ws.cell(row=sub_row, column=COL_QTY).value   = f"=SUM(L{PRODUCT_START_ROW}:L{last_prod})"
    ws.cell(row=sub_row, column=COL_TOTAL).value = f"=SUM(M{PRODUCT_START_ROW}:M{last_prod})"
    ws.cell(row=total_row, column=COL_TOTAL).value = (
        f"=ROUND(SUM(O{PRODUCT_START_ROW}:O{last_prod})+M{other_row},0)"
    )

    # ── 5. Save and return ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
